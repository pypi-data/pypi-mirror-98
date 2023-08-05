# -*- coding: utf-8; -*-
################################################################################
#
#  Rattail -- Retail Software Framework
#  Copyright © 2010-2020 Lance Edgar
#
#  This file is part of Rattail.
#
#  Rattail is free software: you can redistribute it and/or modify it under the
#  terms of the GNU General Public License as published by the Free Software
#  Foundation, either version 3 of the License, or (at your option) any later
#  version.
#
#  Rattail is distributed in the hope that it will be useful, but WITHOUT ANY
#  WARRANTY; without even the implied warranty of MERCHANTABILITY or FITNESS
#  FOR A PARTICULAR PURPOSE.  See the GNU General Public License for more
#  details.
#
#  You should have received a copy of the GNU General Public License along with
#  Rattail.  If not, see <http://www.gnu.org/licenses/>.
#
################################################################################
"""
Excel utilities
"""

from __future__ import unicode_literals, absolute_import

import datetime

import six
import xlrd
from xlrd.xldate import xldate_as_tuple

import openpyxl
from openpyxl.styles import Font, PatternFill
from openpyxl.cell.cell import get_column_letter

from rattail.util import progress_loop


class ExcelReaderXLS(object):
    """
    Basic class for reading Excel "legacy" (.xls) files.

    Uses the ``xlrd`` package to read the files.
    """

    def __init__(self, path, sheet=0, sheet_name=None, header=0,
                 first_data_row=None,
                 datefmt='%Y-%m-%d', strip_fieldnames=True):
        """
        Constructor; opens an Excel file for reading.

        :param header: Which row should be used as the header, i.e. to
           determine field (column) names.  This is a zero-based index, so is 0
           by default (i.e. the first row).

        :param first_data_row: Which is the first row to contain data.  If not
           specified, it will be assumed that data rows begin immediately after
           the header row, as defined by :param:`header`.  This again is
           zero-based, so if the very first row is the true header, but then
           there is another "header" row also, you might specify a value of
           ``2`` here, since the 3rd row is the first to contain data.
        """
        self.book = xlrd.open_workbook(path)
        if sheet_name is not None:
            self.sheet = self.book.sheet_by_name(sheet_name)
        else:
            self.sheet = self.book.sheet_by_index(sheet)
        self.header = header
        if first_data_row is not None:
            self.first_data_row = first_data_row
        else:
            self.first_data_row = self.header + 1
        self.fields = self.sheet.row_values(self.header)
        if strip_fieldnames:
            self.fields = [field.strip() for field in self.fields]
        self.datefmt = datefmt

    def sheet_by_name(self, name):
        return self.book.sheet_by_name(name)

    def read_rows(self, progress=None):
        rows = []

        def append(row, i):
            values = self.sheet.row_values(row)
            data = dict([(self.fields[j], value)
                         for j, value in enumerate(values)])
            rows.append(data)

        progress_loop(append, range(self.first_data_row, self.sheet.nrows), progress,
                      message="Reading data from Excel file")
        return rows

    def parse_date(self, value, fmt=None):
        if isinstance(value, float):
            args = xldate_as_tuple(value, self.book.datemode)
            return datetime.datetime(*args).date()
        if value:
            return datetime.datetime.strptime(value, fmt or self.datefmt).date()

# TODO: this should become a base class and/or wrapper of some sort. for now
# the "default" reader assumes XLS since apparently that is all we have ever
# supported until now...which was a surprise to me...
ExcelReader = ExcelReaderXLS


class ExcelReaderXLSX(object):
    """
    Basic class for reading Excel 2010 (.xslx) files.

    Uses the ``openpyxl`` package to read the files.
    """

    def __init__(self, path, header_row=1, **kwargs):
        """
        Constructor; opens an Excel file for reading.

        :param path: Path to the Excel file.

        :param header_row: Which row contains the column headers.  This is
           1-based, so the 1 is the default (i.e. the first row).
        """
        self.book = openpyxl.load_workbook(filename=path)
        self.sheet = self.book.active

        self.header_row = header_row

        self.fields = None
        # TODO: this seems like a hacky way to get the header fields?  we are
        # "iterating" over the single header row, effectively...
        for row in self.sheet.iter_rows(min_row=self.header_row,
                                        max_row=self.header_row,
                                        values_only=True):
            assert self.fields is None
            self.fields = list(row)

    def read_rows(self, progress=None):
        rows = []

        def append(row, i):
            data = dict([(self.fields[j], value)
                         for j, value in enumerate(row)])
            rows.append(data)

        xlrows = list(self.sheet.iter_rows(min_row=self.header_row + 1,
                                           values_only=True))
        progress_loop(append, xlrows, progress,
                      message="Reading data from Excel file")
        return rows


class ExcelWriter(object):
    """
    Base class for Excel writers.
    """

    def __init__(self, path, fields, sheet_title=None, number_formats={}):
        """
        Constructor; opens an Excel workbook for writing.
        """
        self.path = path
        self.fields = fields
        self.book = openpyxl.Workbook()
        self.sheet = self.book.active
        if sheet_title:
            self.sheet.title = sheet_title
        self.number_formats = number_formats

    def create_sheet(self, title):
        """
        Create a new sheet in the workbook, and make it active.
        """
        self.sheet = self.book.create_sheet(title)
        return self.sheet

    def enable_grid_lines(self, sheet=None, enabled=True):
        if not sheet:
            sheet = self.sheet
        sheet.sheet_view.showGridLines = enabled

    def disable_grid_lines(self, sheet=None):
        self.enable_grid_lines(sheet=sheet, enabled=False)

    def write_header(self, labels=None):
        font = Font(bold=True)
        for i, field in enumerate(self.fields, 1):
            value = field
            if labels and field in labels:
                value = labels[field]
            cell = self.sheet.cell(row=1, column=i, value=value)
            cell.font = font

    def write_row(self, data, row=None):
        """
        Write (append) a single data row to the current sheet.

        :param row: The 1-based row number to which data should be written.
        """
        if row is None:
            raise NotImplementedError("should be able to detect 'next' row here?")

        self.sheet.append(data)

        # apply number formats
        if self.number_formats:
            for col, field in enumerate(self.fields, 1):
                if field in self.number_formats:
                    cell = self.sheet.cell(row=row, column=col)
                    cell.number_format = self.number_formats[field]

        # apply row highlighting
        if row % 2 == 0:
            fill_even = PatternFill(patternType='solid',
                                    fgColor='d9d9d9',
                                    bgColor='d9d9d9')
            for col, field in enumerate(self.fields, 1):
                cell = self.sheet.cell(row=row, column=col)
                cell.fill = fill_even

    def write_rows(self, rows, progress=None):
        """
        Write (append) a sequence of data rows to the current sheet.
        """
        def write(data, i):
            # must add 1 to account for header
            self.write_row(data, row=i + 1)

        progress_loop(write, rows, progress,
                      message="Generating cells for spreadsheet")

    def auto_freeze(self, row=2, column=1):
        """
        Freeze sheet per "the usual"
        """
        self.sheet.freeze_panes = self.sheet.cell(row=row, column=column)

    def auto_filter(self):
        """
        Add auto filters for all columns.
        """
        first = self.sheet.cell(row=1, column=1)
        last = self.sheet.cell(row=self.sheet.max_row, column=self.sheet.max_column)
        cellrange = '{}:{}'.format(first.coordinate, last.coordinate)
        self.sheet.auto_filter.ref = cellrange

    def auto_resize(self, progress=None):
        """
        (Try to) Auto-resize all data columns.
        """
        # note, some of the below uses efficiency tricks from these docs
        # https://openpyxl.readthedocs.io/en/stable/tutorial.html#accessing-many-cells

        # we must calculate desired column widths.  but for sake of progress,
        # we'll iterate through rows instead of columns, to do that.  (this is
        # just to give a higher total for the progress bar.)  so the first pass
        # is really just to cache all existing string widths.
        cached = []

        def cache(row, i):
            cached.append([len(six.text_type(value))
                           for value in row])

        progress_loop(cache, list(self.sheet.values), progress,
                      message="Calculating all string widths")

        # okay, now can determine ideal widths
        column_widths = []

        def calculate(col, i):
            width = max([row_widths[col] for row_widths in cached])
            column_widths.append(width or 5)

        progress_loop(calculate, range(self.sheet.max_column), progress,
                      message="Calculating desired column widths")

        # resize columns
        for i, width in enumerate(column_widths, 1):
            self.sheet.column_dimensions[get_column_letter(i)].width = width + 3

    def save(self, progress=None):
        """
        Save the Excel workbook to file.  If ``progress`` is provided, it will
        be used in a hacky sort of way, i.e. from 0 to 1 only since we have no
        way of knowing true progress for the save operation.  (But it can still
        be nice to let user know this is the step we're on at least.)
        """
        def save(x, i):
            self.book.save(self.path)

        progress_loop(save, range(1), progress,
                      message="Saving workbook to file")
