from tempfile import NamedTemporaryFile

from openpyxl import Workbook, load_workbook
from openpyxl.styles import Font, Alignment

"""
openpyxl不支持旧的.xls文件格式，请使用xlrd读取该文件，或将其转换为最新的.xlsx文件格式。
"""


class OpenpyxlUtils(object):
    def __init__(self):
        self.wb = Workbook()  # 该参数guess_types=True 启用了类容和格式推断但为什么Workbook 没有这个参数
        self.ws = None

    def xls_to_xlsx(self, xlsPath):
        """
        https://blog.csdn.net/test_sir_cao/article/details/79133617
        mime  http://www.w3school.com.cn/media/media_mimeref.asp
        python -m pip install pypiwin32

        '类 Workbooks 的 Open 方法无效
        https://blog.csdn.net/xiaozhuzhu1/article/details/48493529
        出现这个错误的原因是，已有的Excel文件中有无法识别的内容，手动打开的时候，会提示：Excel发现不可读取内容，只要确认打开，再保存一下，再用程序打开的时候就不会出现这个问题了。
        有以下几种：
        51 = xlOpenXMLWorkbook (without macro's in 2007-2016, xlsx) 保存为xlsx
        52 = xlOpenXMLWorkbookMacroEnabled (with or without macro's in 2007-2016, xlsm)
        保存为xlsm带宏的格式
        50 = xlExcel12 (Excel Binary Workbook in 2007-2016 with or without macro's, xlsb)
        以二进制保存的工作表
        56 = xlExcel8 (97-2003 format in Excel 2007-2016, xls)
        以xls，2003格式保存的工作表
        :param xlsPath:
        :return:
        """
        import win32com.client
        excel = win32com.client.gencache.EnsureDispatch('Excel.Application')  # 要看MIME手册
        wb = excel.Workbooks.Open(xlsPath)
        wb.SaveAs(xlsPath + 'x', FileFormat=51)
        wb.Close()
        excel.Application.Quit()

    def set_wb_only_read(self, filename, read_only=False):
        """
        只读方式打开excel
        :param filename:
        :param read_only:
        :return:
        """
        self.wb = load_workbook(filename=filename, read_only=read_only)
        return self.wb

    def set_wb_only_write(self):
        """
        按照写的方式打开一个新的wb
        :return:
        """
        self.wb = Workbook(write_only=True)
        return self.wb

    def create_ws(self, sheetname='', sheetnum=0):
        """
        获取 ws sheet 对象 此函数使用_active_sheet_index属性，默认设置为0。
       除非您修改其值，否则您将始终使用此方法获取第一个工作表。
        :return:
        """
        if sheetname:
            self.ws = self.wb.create_sheet(sheetname, sheetnum)
        else:
            # 获取第一个 sheet
            self.ws = self.wb.active
        return self.ws

    def set_ws_title(self, title):
        """
        设置工作表的名字
        :param title:
        :return:
        """
        self.ws.title = title

    def get_ws(self, title):
        """
        获取一个sheet 通过title
        :param title:
        :return:
        """
        self.ws = self.wb[title]
        return self.ws

    def get_list_sheet(self):
        """
        可以通过 wb.sheetnames 获取列表
        :return:
        """
        for sheet in self.wb:
            yield sheet

    def get_sheet_copy(self):
        """
        得到当前sheet的一个副本
        :return:
        """
        return self.wb.copy_worksheet(self.ws)

    def get_ws_value(self, key):
        """
        如 A4
        :param key:
        :return:
        """
        return self.ws[key].value

    def set_ws_value(self, key, value):
        self.ws[key] = value
        return self.ws[key]

    def get_sheet_value_cell(self, row, colum, value):
        """
        使用cell方法访问指定单元格如果没有会创建
        >>> for i in range(1,101):
        ...        for j in range(1,101):
        ...            ws.cell(row=i, column=j)
        将在内存中创建100x100个单元格，无需任何操作。
        :param row:
        :param colum:
        :param value:
        :return:
        """
        return self.ws.cell(row, colum, value)

    def get_sheet_value_slicing(self, index1, index2=None):
        """

        >>> colC = ws['C']
        >>> col_range = ws['C:D']
        >>> row10 = ws[10]
        >>> row_range = ws[5:10]

        :param index1: 数字或字母
        :param index2:
        :return: 通过传入的参数获取多个 tuple结果 比如 传入 A C 将返回A B C 3个tuple元组
        也可传入数字 传入数字将返回行，传入字母将返回列 也是以元组的方式返回
        注意 迭代元组得到 cell对象 可以 通过 cell.value得到值cell.row，cell.column得到行和列
        """
        if index2 is None:
            return self.ws[index1]
        return self.ws[index1:index2]

    def get_sheet_iter_rows(self, min_row, max_col, max_row):
        for row in self.ws.iter_rows(min_row, max_col, max_row):
            for cell in row:
                yield cell

    def get_sheet_iter_cols(self, min_row, max_col, max_row):
        for row in self.ws.iter_cols(min_row, max_col, max_row):
            for cell in row:
                yield cell

    def get_sheet_all_rows(self, is_row=True):
        """
        得到所有行和列 使用 tuple()查看对象
        :return:
        """
        if is_row:
            return self.ws.rows
        else:
            return self.ws.columns

    def set_sheet_cell_value(self, cell, value):
        """
        为单元格设置一个值
        :param cell:
        :param value:
        :return:
        """
        cell.value = value
        print(cell.value)

    def save_wb(self, excelname):
        """
        保存为excel文件
        :param excelname:
        :return:
        """
        self.wb.save(excelname)

    def save_stream(self):
        """
        保存到流
        :return:
        """
        with NamedTemporaryFile() as tmp:
            self.wb.save(tmp.name)
            tmp.seek(0)
            stream = tmp.read()
        return stream

    def save_wb_template(self, excelname, templatename):
        """
        >>> wb = load_workbook('document.xlsx')
        >>> wb.template = True
        >>> wb.save('document_template.xltx')
        您可以指定属性template = True，以将工作簿另存为模板：
        或者将此属性设置为False（默认），以保存为文档
        >>> wb = load_workbook('document_template.xltx')
        >>> wb.template = False
        >>> wb.save('document.xlsx', as_template=False)
        :param excelname:
        :return:
        """

        self.wb.template = True
        self.wb.save(excelname)

    def max_column(self):
        """
        所有列
        :return:
        """
        return self.ws.max_column

    def max_row(self):
        """
        所有行
        :return:
        """
        return self.ws.max_row

    def save_excel(self, filename):
        """
        保存excel文件
        :param filename:
        :return:
        """
        self.wb.save(filename)

    def get_font(self):
        """
        name 字符串 字体名称，例如'Calibri'或'Times New Roman'

        size 整数 字体大小

        bold 布尔 True，粗体字体

        italic 布尔 True，斜体字体
        :return:
        """
        fontObj1 = Font(bold=True)
        return fontObj1

    def get_Alignment(self):
        al = Alignment(horizontal="center")
        return al

    def set_font(self, fontobj, alignment, cells=[]):
        """
        :param cell:  A1
        :param fontobj:
        :return:
        """
        for cell in cells:
            self.ws[cell].font = fontobj
            self.ws[cell].alignment = alignment
