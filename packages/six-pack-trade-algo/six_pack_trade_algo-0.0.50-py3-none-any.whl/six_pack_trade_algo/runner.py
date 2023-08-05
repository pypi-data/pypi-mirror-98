import pandas as pd
from ta.utils import dropna
from datetime import datetime as dt, timedelta
import ta
import time
import argparse
from multiprocessing import Process, Manager
import threading
from openpyxl import load_workbook, Workbook
import configparser
import pathlib
'''============================================================================'''
##################################################################################
'''




'''
##################################################################################
'''============================================================================'''
try:
    from order import *
    from algos import *
    from util import log as log
    from util import DataSource
    from util import Stream
    print("runner.py dev")
except Exception as e:
    from six_pack_trade_algo import *
    # print("runner.py {}".format(e))
    
    
def MainAlgoRunner():
    # Load configs according to comand line input
    parser = argparse.ArgumentParser()
    parser.add_argument("--main_config", default="DEFAULT",type=str)
    parser.add_argument('--main_path', default="./",type=str)
    parser.add_argument('--strategy_config', default=["SHORT_PAPER_1.0"], nargs='+')
    args = parser.parse_args()
    if args.main_path == "./" or args.main_path == "":
        main_path = str(pathlib.Path(__file__).parent.absolute())
    else:
        main_path = args.main_path

    config_list = []
    try:
        try:
            main_config = configparser.ConfigParser()
            main_config.read(main_path)
            main_config = main_config[args.main_config]
        except:
            raise Exception("Cannot find main config")
            
        test           = main_config["test"]
        data_path      = main_config["data_path"]
        data_source    = main_config["data_source"]
        api_key_id     = main_config["api_key_id"]
        api_secret_key = main_config["api_secret_key"]
        api_base_url   = main_config["api_base_url"]
        try:
            start_date = main_config["start_date"]
            end_date = main_config["end_date"]
        except Exception as e:
            print("************ No dates found in config for Backtester. Using defaults. ************        ")
            start_date = "2020-06-01"
            end_date = "2020-06-15"


        
        str_config_parser = configparser.ConfigParser()
        for profile in args.strategy_config:
            str_config_parser.read(data_path + "configs/str.cnt")
            str_config = str_config_parser[profile]
            config_list.append(str_config)
        
    except Exception as e:
        print("************ Error loading configs in main. ************        " + str(e))
        return 1

    # 
    # make the factory
    # 
    algoFactory = AlgoFactory()
    orderManagerFactory = OrderManagerFactory(test=test)

    try:
        master_wb = load_workbook(filename=data_path + "master.xlsx")
    except Exception as e:
        print("Error opening master excel file. {}...\nTrying to create new workbook.".format(e))
        new_wb = Workbook()
        new_wb.save(data_path + "master.xlsx")
        new_wb.close()
        master_wb = load_workbook(filename=data_path + "master.xlsx")
        
    
    algo_list = {}
    thread_list = []
    sheet_dict = {}

    return_dict = Manager().dict()
    sm='shared_mem_prefix'
    stream = Stream(
                key_id=api_key_id ,secret_key=api_secret_key,
                base_url=api_base_url, data_path=data_path,
                sm=sm, test=test
    )
    stream_thread = threading.Thread(target=stream.ws_start, daemon=True)
    stream_thread.start()

    i = 2
    sheet_dict = {}
    for config in config_list:
        guid = config['guid']
        data_source = DataSource(
                    key_id=api_key_id ,secret_key=api_secret_key ,
                    base_url=api_base_url, data_source=data_source, tickers=config["tickers"].split(","),
                    start_date=start_date ,end_date=end_date
                    )
        print("Setup datasource for {}".format(guid))

        orderManagerFactory.setType(config['type'])
        order_manager = orderManagerFactory.build(
                                GUID=guid,
                                data_path=data_path,
                                data_source=data_source,
                                order_config=config,
                                sm=sm
                                )
        print("Setup order manager for {}".format(guid))

        algoFactory.setType(config['type'])
        algo = algoFactory.build(order_manager=order_manager,
                                 data_path=data_path,
                                 data_source=data_source,
                                 GUID=guid,
                                 config=config
                                )
        print("Setup algo for {}".format(guid))

        
        # Start the process and set the name to be the guid. Each guid will always have its own process/thread
        # you can think of return_dict as shared memory that allows us to get certain results from threads
        # after we close them.
        algo_thread = Process(target=algo.run, name=guid, args=(return_dict,))
        algo_thread.start()
        print("Started Process for {}".format(guid))
        
        
        algo_list[guid] = algo
        thread_list.append((guid ,algo_thread))
        
        
        share_val = algo.order_manager.get_total_value()
        # Find sheet with GUID, create if DNE
        try:
            sheet = master_wb[guid]
        except:
            master_wb.create_sheet(guid)
            master_wb.save(data_path + "master.xlsx")
            master_wb.close()
            master_wb = load_workbook(filename=data_path + "master.xlsx")
            sheet = master_wb[guid]
            sheet.cell(row=1, column=1, value="Date")
            sheet.cell(row=1, column=2, value="Wallet")
            sheet.cell(row=1, column=3, value="Tickers")
            sheet.cell(row=1, column=4, value="Tick Period")
            sheet.cell(row=1, column=5, value="Algo Details")
            sheet.cell(row=1, column=6, value="OrderManager Details")
            sheet.cell(row=1, column=7, value="Share Value")
            sheet.cell(row=1, column=8, value="Extra 1")
            sheet.cell(row=1, column=9, value="Extra 2")

        #Find first available row
        d = 2
        while True:
            try:
                sheet[d][0] = ""
                d += 1
            except:
                sheet = master_wb[guid]
                break
        sheet.cell(row=d, column=1, value=dt.today().strftime('%Y-%m-%d'))
        sheet.cell(row=d, column=2, value=order_manager.wallet)
        sheet.cell(row=d, column=3, value=', '.join(order_manager.tickers))
        sheet.cell(row=d, column=4, value=algo.tick_period)
        sheet.cell(row=d, column=5, value=algo.print_details())
        sheet.cell(row=d, column=6, value=order_manager.print_details())
        sheet.cell(row=d, column=7, value=share_val)
        sheet_dict[guid] = d
        i+=1
        print("Finished init for {}".format(guid))
        

    # Check if threads are done and save out
    while 0 < len(thread_list):
        remove_list = []
        for guid, thread in thread_list:
            if not thread.is_alive():
                thread.join()
                print("JOIN THREAD")
                algo_finished = algo_list[guid]
                ret = return_dict[guid]
                pos = ret[1]
                total_shares = 0
                for shares in list(pos.values()):
                    total_shares += shares[0]
                share_val = 0
                for p in pos:
                    share_val += pos[p][0] * algo_finished.data_source.get_last_trade(p)["price"]
                sheet = master_wb[guid]
                try:
                    sheet_dict[sheet] += 1
                    
                except:
                    sheet_dict[sheet] = 1
                d = sheet_dict[sheet]

                sheet.cell(row=d, column=8, value=ret[0])
                sheet.cell(row=d, column=9, value=total_shares)
                sheet.cell(row=d, column=10, value=share_val)
                sheet.cell(row=d, column=11, value=share_val + ret[0])
                master_saved = False
                while not master_saved:
                    try:
                        master_wb.save(data_path + "master.xlsx")
                        master_saved = True
                    except Exception as e:
                        print(e)
                        print("CANNOT SAVE TO MASTER")
                        time.sleep(1)
                remove_list.append((guid,thread))
        for tup in remove_list:
            thread_list.remove(tup)
        time.sleep(1)
            
            
    while algo_finished.data_source.get_clock()["is_open"] and stream.get_num_open() != 0:
        time.sleep(5)
        print("Keeping process alive for stream thread")
    return 0



if __name__ == "__main__":
    ret =  MainAlgoRunner()
    if not (ret == 0):
        print("ERROR IN MAIN: {}".format(ret))
    else:
        print("SUCCESS")
    exit()

