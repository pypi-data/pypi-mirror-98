import pandas as pd
from ta.utils import dropna
from datetime import datetime, timedelta
import ta
import time
import argparse
from multiprocessing import Process, Manager
from openpyxl import load_workbook


from algotrader.algos import AlgoRsi as Algo
from algotrader.util import algoLogger.log as log
from algotrader.util import DataSource

# import alpaca_trade_api as tradeapi
# import pandas as pd
# import time
# from ta.volatility import BollingerBands
# from ta import add_all_ta_features
# from ta.utils import dropna
# from datetime import datetime, timedelta
# import ta
# import time
# import matplotlib.pyplot as plt
# import argparse
# # from threading import Thread
# from multiprocessing import Process, Manager
# from algoBb import AlgoBb as Algo
# from openpyxl import load_workbook
# import copy
# 
# from data_source import DataSource


FILE_ALGO_TYPE = "BB"

def main():
    parser = argparse.ArgumentParser()
    parser.add_argument('--test_mode', default="False", type=str)
    parser.add_argument('--data_source_mode', default="paper", type=str)    
    args = parser.parse_args()
    
    datasource = DataSource(args.data_source_mode)

    # Iterate through master csv
    #   skip first line
    # create a thread for each line in the master csv
    #   This will involve creating algo objects for each thread
    #   This will involve creating a thread object
    
    master_wb = load_workbook(filename="./master.xlsx")
    sheet = master_wb[FILE_ALGO_TYPE]

    return_dict = Manager().dict()

    algo_list = {}
    thread_list = []
    i = 0
    while i < int(sheet.max_row):
        i += 1
        if (sheet[i][0].value == None or sheet[i][0].value == ""):                                           # Column with GUID
            break
        if (FILE_ALGO_TYPE != "" and FILE_ALGO_TYPE not in sheet[i][0].value):  # If not in main.py set FILE_ALGO_TYPE != ""
            continue

        tickers = sheet[i][2].value.split(",")
        # exit()
        algo = Algo(sheet[i][0].value, datasource, sheet[i][1].value, tickers, sheet[i][3].value, sheet[i][4].value, sheet[i][5].value, sheet[i][6].value )
        
        # check if the algo has prev day values in the spread sheet
        # DISABLED FOR NOW
        # if FILE_ALGO_TYPE != "test":
        #     try:
        #         if sheet[i][7].value != None and :
        #             algo.wallet = sheet[i][7].value
        #         else:
        #             raise Exception("None wallet found in csv or test")
        #     except Exception as e:
        #         print(e)
        #         print("Can't find prev day data for algo: " + str(sheet[i][0].value))
        
        if args.test_mode == "False":
            algo.test_mode = False
        else:
            algo.test_mode = True
            
        algo_thread = Process(target=algo.run, name=str(i), args=(return_dict,))
        algo_thread.start()
        
        algo_list[i] = algo
        thread_list.append(algo_thread)
    while 0 < len(thread_list):
        remove_list = []
        for thread in thread_list:
            if not thread.is_alive():
                thread.join()
                i = int(thread.name)
                algo_finished = algo_list[i]
                ret = return_dict[algo_finished.GUID]
                sheet[i][7].value = ret[0]
                pos = ret[1]
                total_shares = 0
                for shares in list(pos.values()):
                    total_shares += shares
                sheet[i][8].value = total_shares
                share_val = 0
                for p in pos:
                    share_val += pos[p] * algo_finished._get_last_trade(p)
                sheet[i][9].value = share_val
                sheet[i][10].value = share_val + algo_finished.wallet            
                master_wb.save("./master.xlsx")
                remove_list.append(thread)
        for thread in remove_list:
            thread_list.remove(thread)
        time.sleep(1)
    
    
if __name__ == "__main__":
    main()
    exit()
    

