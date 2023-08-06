# -*- coding: utf-8 -*-
# @Time    : 2021/3/10 1:31 下午
# @Author  : john
# @Site    :
# @File    : __init__.py.py
# @Software: PyCharm
name = "pcacode"
from openpyxl import load_workbook
import os


class pcacode:

    def __init__(self,
                 file_page=os.path.abspath(os.path.dirname(__file__))+'/data/2020年11月中华人民共和国县以上行政区划代码.xlsx',
                 nations_file=os.path.abspath(os.path.dirname(__file__))+'/data/56_nations.xlsx',
                 switch_data={},
                 dict_data={}):
        # 1 将地区吗加载成文件

        self.__file_page = file_page
        self.__data = self.__load_data(file_page=self.__file_page)
        self.__nations_file = nations_file
        self.__nations_data = self.__load_nations_data(file_page=self.__nations_file)
        self.__switch_data = switch_data  # 增加省名替换
        self.__dict_data = dict_data  # 更新地区码数据
        self.__up_load_data()

    # 加载数据
    def __load_nations_data(self, file_page):
        """
        加载56个名族数据
        :param file_page: 文件路径
        :return: 返回包含全部名族的列表 ['汉族', '蒙古族', '回族']
        """
        workbook = load_workbook(file_page)
        booksheet = workbook.active
        rows = booksheet.rows
        data = []
        for row in rows:
            data.append(row[0].value)
        return data

    # 更新数据 （默认增加直辖市原始数据中不存在）
    def __up_load_data(self):
        """
        行政区划代码 更新 用于小范围内变更数据
        :param
        :return:
        """
        base_dict = {'110100': '直辖市', '310100': '直辖市', '120100': '直辖市', '500100': '直辖市'}  # 默认数据 原数据中不存在
        base_dict.update(self.__dict_data)
        for key, value in base_dict.items():
            self.__data[key] = value

    # 加载民族数据
    def __load_data(self, file_page):
        """
        :param file_page:  文件路径
        :return: 返回列表 {'地区码1'：'名称1'，'地区码2'：'名称2'………………}
        """
        workbook = load_workbook(file_page)
        booksheet = workbook.active
        rows = booksheet.rows
        data = {}
        for row in rows:
            data[str(row[0].value)] = str(row[1].value)

        return data

    # 常见省名替换
    def __switch_address(self, province):
        item = {
            '西藏': '西藏自治区',
            '新疆': '新疆维吾尔自治区',
            '广西': '广西壮族自治区',
            '内蒙': '内蒙古自治区',
            '内蒙古': '内蒙古自治区',
            '宁夏': '宁夏回族自治区',
            '北京': '北京市',
            '天津': '天津市',
            '上海': '上海市',
            '重庆': '重庆市',
        }
        item.update(self.__switch_data)
        if province in item:
            return item[province]
        else:
            if '省' in province:
                return province
            else:
                if '市' in province:
                    return province
                return province + '省'

    # 根据地区码找城市名称
    def __fuzzy_finder_city(self, key):
        """
        根据地区码查找 城市名称
        :param key:
        :return:
        """
        key_len = len(key)

        area_dict = {}
        for k, v in self.__data.items():
            if k[:key_len] == key and k[4:] == '00' and k[2:] != '0000':
                area_dict[k] = v

        # print(area_dict)
        return area_dict

    # 根据地区吗找区域名称
    def __fuzzy_finder_area(self, key):
        """
        根据地区码查找 区县名称
        :param key:
        :return:
        """
        area_dict = {}
        for k, v in self.__data.items():
            if k[:4] == key[:4] and k[4:] != '00':  # 省开头 非空结尾
                area_dict[k] = v

        # print(area_dict)
        return area_dict

    # 获取省名称
    def __get_province_code(self, province):

        z = [k for k, v in self.__data.items() if v == province]  # 可能多个 潜在问题 待优化处理
        if z != []:
            return z[0]
        return None

    # 模糊化城市名称
    def __fuzzy_city_name(self, city_name):
        """
        模糊名称
        :param city_name: 输入的城市名称
        :return:   返回模糊后的名称
        """
        # sort_name = self.nations_data + ['市', '州', '自治州', '自治县', '市', '地区']
        # for rm_name in sort_name:
        #     name = city.strip(rm_name)  # 简称会省略区县自治区自治县要剔除
        #     if len(name) >= 2:
        #         city = name

        return city_name

    # 模糊化区县名称
    def __fuzzy_area_name(self, area_name):
        """
        模糊名称
        :param area_name: 输入的区县名称
        :return:   返回模糊后的名称
        """
        return area_name

    # 获取详细地区码（处理的核心）
    def __get_area_code(self, province, content=''):

        # province=['省','市','自治区'] #备用替换后缀防呆

        # 1 获取省码  （省名的简称）
        province_name = self.__switch_address(province=province)  # 替换省名称
        province_code = self.__get_province_code(province=province_name)  # 获取省级的地区码

        if not province_code:
            print('省名错误 province Erreo')
            return []

        # 2 获取市码
        city_dict = self.__fuzzy_finder_city(province_code[:2])  # 根据省码反查市码
        # print(city_dict)

        accurate_city_dict = {}

        for code, city in city_dict.items():  # 循环查找市区

            if city in content:
                accurate_city_dict[code] = city
                continue
            elif self.__fuzzy_city_name(city) in content:
                accurate_city_dict[code] = city
                continue

        city_flag = 0
        if accurate_city_dict == {}:  # 没有查找到市区信息
            city_flag = 1
            accurate_city_dict = city_dict  # 就用全部城市进入下一轮

        # 3 通过市码获取区码
        # 地址列表
        address_code = []
        # 城市列表
        city_code_list = []
        for code, city in accurate_city_dict.items():  # 循环精确的市列表
            area_dict = self.__fuzzy_finder_area(code)  # 通过市列表推倒区域
            area_address_code = []
            for area_code, area_name in area_dict.items():
                if area_name in content:
                    area_address_code.append(area_code)
                    address_code += area_address_code
                    continue
                elif self.__fuzzy_area_name(area_name) in content:
                    area_address_code.append(area_code)
                    address_code += area_address_code
                    continue
                if city_flag == 0 and code not in city_code_list:
                    city_code_list.append(code)

        # 地址列表        address_code
        # 城市列表        city_code_list
        # 将 city_code_list 并入 address_code

        for city_code in city_code_list:
            tag = 0
            for code in address_code:
                if city_code[:4] == code[:4]:
                    tag = 1
            if tag == 0:
                address_code.append(city_code)

        if not address_code and city_flag == 0:

            #   如果区域为空 而且找到城市 返回城市列表
            for area_code, area_name in accurate_city_dict.items():
                # print(area_code,area_name)
                address_code.append(area_code)

            # print(address_code)
            return address_code

        #   如果区域为空 没有找到城市 返回省
        if not address_code and city_flag == 1:
            # print([province_code])
            return [province_code]

        address_code = list(set(address_code))
        # print(address_code)
        return address_code

    # 根据地区吗生成城市
    def __get_area_info(self, address_code):
        item = []
        for code in address_code:
            # print((code[:2] + '0000'))
            # print(self.data[code[:2] + '0000'])
            item.append((self.__data[(code[:2] + '0000')], self.__data[(code[:4] + '00')], self.__data[code], code))
        return item

    # 获取信息
    def get_info(self, province, content):
        return self.__get_area_info(self.__get_area_code(province=province, content=content))

