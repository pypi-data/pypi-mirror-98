import re
from functools import partial
from itertools import islice
from itertools import chain
from operator import itemgetter

from zope.interface import Invalid
from zope.component import getUtility
from zope.schema.interfaces import IVocabularyFactory

import Acquisition

from Products.CMFPlone.utils import safe_unicode
from Products.Five.browser import BrowserView
from Products.statusmessages.interfaces import IStatusMessage

from plone import api

import openpyxl

from emrt.necd.content.vocabularies import get_registry_interface_field_data
from emrt.necd.content.vocabularies import INECDVocabularies
from emrt.necd.content.observation import create_comment
from emrt.necd.content.observation import inventory_year
from emrt.necd.content.question import create_question


PORTAL_TYPE = 'Observation'

UNUSED_FIELDS = ['closing_comments', 'closing_deny_comments']

UNCOMPLETED_ERR = u'The observation on row no. {} seems to be a bit off. ' \
                  u'Please fill all the fields as shown in the import file' \
                  u' sample. '

WRONG_DATA_ERR = u'The information you entered in the {} section ' \
                 u'of row no. {} is not correct. Please consult the columns' \
                 u' in the sample xls file to see the correct set of data.' \

DONE_MSG = u'Successfully imported {} observations!'

PROJECTION_COLS = (
    'nfr_code_inventory', 'year', 'reference_year',
    'pollutants', 'scenario', 'review_year',
    'activity_data_type', 'activity_data', 'ms_key_category',
    'parameter', 'highlight'
)

INVENTORY_COLS = (
    'year', 'pollutants', 'review_year', 'fuel',
    'ms_key_category', 'parameter', 'highlight'
)


def _read_row(idx, row):
    val = itemgetter(idx)(row).value

    if not val:
        return ''

    if isinstance(val, (int, long)):
        val = safe_unicode(str(val))
    return val.strip()


def _multi_rows(row):
    splitted = re.split(r'[,\n]\s*', row)
    return tuple(val.strip() for val in splitted)


def _get_vocabulary(context, name):
    factory = getUtility(IVocabularyFactory, name=name)
    return factory(context)


def _cached_get_vocabulary(context, prefix=''):
    cache = {}

    def get_vocabulary(name):
        name = prefix + name if prefix else name
        if name not in cache:
            cache[name] = _get_vocabulary(context, name)
        return cache[name]

    return get_vocabulary


def get_constants(context):
    XLS_COLS = {}
    XLS_COLS['text'] = partial(_read_row, 0)
    XLS_COLS['country'] = partial(_read_row, 1)
    XLS_COLS['nfr_code'] = partial(_read_row, 2)
    next_col = 3

    # choose columns based on context
    iter_cols = (
        PROJECTION_COLS if context == 'projection'
        else INVENTORY_COLS
    )

    # setup readers
    for idx, col in enumerate(iter_cols, start=next_col):
        XLS_COLS[col] = partial(_read_row, idx)

    # last column always the initial Q&A question
    XLS_COLS['question'] = partial(_read_row, idx + 1)

    return XLS_COLS


def find_dict_key(vocabulary, value):
    for key, val in vocabulary.items():
        if isinstance(val, list):
            if value in val:
                return key
        elif isinstance(val, Acquisition.ImplicitAcquisitionWrapper):
            if val.title == value:
                return key
        elif val == value:
            return key

    return False


def key_in_vocab(vocab, key):
    for term in vocab:
        if key == term.title:
            return term.token
    return False


def error_status_messages(context, request, messages):
    status = IStatusMessage(request)
    for message in messages:
        status.addStatusMessage(message, 'error')
    url = context.absolute_url() + '/observation_import_form'
    return request.response.redirect(url)


class Entry(object):

    def __init__(self, row, constants, is_projection, get_vocabulary):
        self.row = row
        self.constants = constants
        self.is_projection = is_projection
        self.get_vocabulary = get_vocabulary

    @property
    def title(self):
        return True

    @property
    def text(self):
        return self.constants['text'](self.row)

    @property
    def country(self):
        country_voc = self.get_vocabulary('eea_member_states')
        cell_value = self.constants['country'](self.row)
        return key_in_vocab(country_voc, cell_value)

    @property
    def nfr_code(self):
        return self.constants['nfr_code'](self.row)

    @property
    def nfr_code_inventory(self):
        nfr = self.constants['nfr_code_inventory'](self.row)
        return nfr if nfr != '' else None

    @property
    def year(self):
        # Projection year
        if self.is_projection:
            years = _multi_rows(self.constants['year'](self.row))
            if years == ('',):
                return ''
            proj_years = [u'2020', u'2025', u'2030', u'2040', u'2050']
            is_correct = bool(set(years) & set(proj_years))
            return list(years) if is_correct else False

        # Inventory year
        cell_value = self.constants['year'](self.row)
        try:
            inventory_year(cell_value)
            return cell_value
        except Invalid:
            return False

    @property
    def reference_year(self):
        return int(self.constants['reference_year'](self.row))

    @property
    def pollutants(self):
        pollutants_voc = self.get_vocabulary('pollutants')
        cell_value = _multi_rows(self.constants['pollutants'](self.row))
        keys = [key_in_vocab(pollutants_voc, key) for key in cell_value]
        if False in keys:
            return False
        return keys

    @property
    def scenario(self):
        scenario_voc = self.get_vocabulary('scenario_type')
        cell_value = _multi_rows(self.constants['scenario'](self.row))
        if cell_value == ('',):
            return None
        keys = [key_in_vocab(scenario_voc, key) for key in cell_value]
        if False in keys:
            return False
        return keys

    @property
    def review_year(self):
        return int(self.constants['review_year'](self.row))

    @property
    def activity_data_type(self):
        activity_voc = self.get_vocabulary('activity_data_type')
        cell_value = self.constants['activity_data_type'](self.row)
        if cell_value == '':
            return None
        if cell_value not in activity_voc:
            return False
        return cell_value

    @property
    def activity_data(self):
        activity_voc = self.get_vocabulary('activity_data')
        cell_value = _multi_rows(self.constants['activity_data'](self.row))
        if cell_value == ('',):
            return None
        elif not self.activity_data_type:
            return False

        keys = [
            key_in_vocab(activity_voc, val)
            for val in cell_value
        ]

        if False in keys:
            return False
        else:
            activity_data_registry = get_registry_interface_field_data(
                INECDVocabularies, 'activity_data')
            if (
                set(activity_data_registry[self.activity_data_type])
                    .difference(cell_value)):
                return False
        return keys

    @property
    def fuel(self):
        fuel_voc = self.get_vocabulary('fuel')
        cell_value = self.constants['fuel'](self.row)
        if cell_value != '':
            return key_in_vocab(fuel_voc, cell_value)
        # This field can be none because it's not manadatory
        return None

    @property
    def ms_key_category(self):
        cell_value = self.constants['ms_key_category'](self.row).title()

        if cell_value == 'True':
            return cell_value
        elif cell_value == '':
            # openpyxl takes False cell values as empty strings so it is easier
            # to assume that an empty cell of the MS Key Category column
            # evaluates to false
            return 'False'

        # For the incorrect data check
        return False

    @property
    def parameter(self):
        parameter_voc = self.get_vocabulary('parameter')
        cell_value = _multi_rows(self.constants['parameter'](self.row))
        keys = [key_in_vocab(parameter_voc, key) for key in cell_value]
        if False in keys:
            return False
        return keys

    @property
    def highlight(self):
        highlight_voc = self.get_vocabulary('highlight')
        col_desc_flags = self.constants['highlight'](self.row)
        if col_desc_flags != '':
            cell_value = _multi_rows(col_desc_flags)
            keys = [key_in_vocab(highlight_voc, key) for key in cell_value]
            if False in keys:
                return False
            else:
                return keys
        else:
            # This field can be none because it's not manadatory
            return None

    @property
    def question(self):
        val = self.constants['question'](self.row)
        if val is not None:
            return val.strip()

    def get_fields(self):
        try:
            # Moving activity_data_type field first to
            # validate activity_data values
            fields = self.constants.keys()
            fields.insert(0, fields.pop(fields.index('activity_data_type')))
        except ValueError:
            pass

        return {
            name: getattr(self, name)
            for name in fields
            if name not in UNUSED_FIELDS
            and name != 'question'
        }


def _create_observation(entry, context, request, portal_type, obj):
    obj.row_nr += 1

    fields = entry.get_fields()

    errors = []

    if '' in fields.values():
        errors.append(UNCOMPLETED_ERR.format(obj.row_nr - 1))
        # return error_status_message(
        #     context, request, UNCOMPLETED_ERR.format(obj.row_nr - 1)
        # )

    elif False in fields.values():
        key = find_dict_key(fields, False)
        key = 'description flags' if key == 'highlight' else key
        errors.append(WRONG_DATA_ERR.format(key, obj.row_nr - 1))

    if errors:
        return None, errors

    # Values must be boolean
    if fields['ms_key_category'] == 'True':
        fields['ms_key_category'] = True
    else:
        fields['ms_key_category'] = False

    content = api.content.create(
        context,
        type=portal_type,
        title=getattr(entry, 'title'),
        **fields
    )

    question_text = entry.question
    if question_text:
        question = create_question(content)
        question.id = 'question-1'
        content[question.id] = question
        create_comment(question_text, content[question.id])

    obj.num_entries += 1

    return content, errors


class ObservationXLSImport(BrowserView):

    num_entries = 0
    row_nr = 2

    def valid_rows_index(self, sheet):
        """There are some cases when deleted rows from an xls file are seen
        as empty rows and the importer tries to create an object with no data
        """
        idx = 1
        for row in sheet:
            if any(cell.value for cell in row):
                idx += 1
        return idx

    def do_import(self):
        xls_file = self.request.get('xls_file', None)

        wb = openpyxl.load_workbook(xls_file, read_only=True, data_only=True)
        sheet = wb.worksheets[0]

        max = self.valid_rows_index(sheet)

        # skip the document header
        valid_rows = islice(sheet, 1, max - 1)

        constants = get_constants(self.context.type)
        get_vocabulary = _cached_get_vocabulary(
            self.context, prefix='emrt.necd.content.')

        entries = []
        for row in valid_rows:
            entries.append(
                Entry(
                    row, constants,
                    self.context.type == 'projection',
                    get_vocabulary
                )
            )

        for entry in entries:
            _, errors = _create_observation(
                entry, self.context, self.request, PORTAL_TYPE, self
            )
            if errors:
                return error_status_messages(
                    self.context, self.request, errors)

        if self.num_entries > 0:
            status = IStatusMessage(self.request)
            status.addStatusMessage(DONE_MSG.format(self.num_entries))

        return self.request.response.redirect(self.context.absolute_url())
