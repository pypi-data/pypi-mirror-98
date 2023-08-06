"""
    Copyright 2019 Simon Vandevelde, Bram Aerts, Joost Vennekens
    This code is licensed under GNU GPLv3 license (see LICENSE) for more
    information.
    This file is part of the cDMN solver.
"""
from typing import List, Dict
import openpyxl
import itertools
import numpy as np
from cdmn.table import Table
from cdmn.glossary import Glossary
from cdmn.idply import Parser
from cdmn.post_process import merge_definitions
import re


def fill_in_merged(file_name: str, sheet_names: List[str] = None) \
                                                -> List[np.array]:
    """
    Loads up a specific sheet and returns it.
    A sheet is comprised of a list of tables.

    :arg file_name: path to the xlsx file.
    :arg sheet: name of the sheet to load in.
    :returns List<np.array>: a list of all the tables in a sheet.
    """
    def rang(m, n): return range(m, n + 1)

    wb = openpyxl.load_workbook(file_name)
    sheets = []
    for sheet_name in sheet_names:
        sheets.append(wb[sheet_name])

    for sheet in sheets:
        ranges = list(sheet.merged_cells.ranges)
        for xyrange in ranges:
            b = xyrange.bounds
            sheet.unmerge_cells(str(xyrange))

            for x, y in itertools.product(rang(b[0], b[2]), rang(b[1], b[3])):
                sheet.cell(y, x).value = sheet.cell(xyrange.min_row,
                                                    xyrange.min_col).value
    return [np.array(list(s.values)) for s in sheets]


def find_first(sheet: np.array) -> tuple:
    """
    TODO

    :arg sheet: the sheet
    """
    for x, y in itertools.product(*[list(range(s)) for s in sheet.shape]):
        if sheet[x, y]:
            return x, y
    return (None, None)


def explore(sheet: List[np.array], boundaries: List[int]):
    """
    Tries to find the ranges for each table.

    :arg sheet: the sheet containing all the tables.
    :arg bounaries: a list containing the theoretic boundaries.
    :returns None:
    """
    startx = boundaries[0]
    endx = boundaries[2]
    starty = boundaries[1]
    endy = boundaries[3]

    # This is a hack to make sure data tables with no input can get found.
    if re.match(r"(?i)DataTable|Data Table", sheet[startx, starty],
                re.IGNORECASE):
        endx += 1
        endy += 1

    while True:
        changed = False

        try:
            while any(sheet[endx+1, starty: endy+1]):
                endx += 1
                changed = True
        except IndexError:
            pass
        try:
            while any(sheet[startx: endx+1, endy+1]):
                endy += 1
                changed = True
        except IndexError:
            pass
        if not changed:
            break
    boundaries[2] = endx + 1
    boundaries[3] = endy + 1


def identify_tables(sheets: List[np.array]) -> List[np.array]:
    """
    Function which looks for all the tables in a given sheet.
    Creates a list of boundaries for the tables.

    :arg sheets: a list containing a numpy array representing the sheet.
    :returns List[np.array]: containing boundaries of all the tables.
    """
    tables = []
    for sheet in sheets:
        while True:
            index = find_first(sheet)
            if index[0] is None and index[1] is None:
                break
            boundaries = list(index + index)
            explore(sheet, boundaries)
            tables.append(sheet[boundaries[0]: boundaries[2],
                          boundaries[1]: boundaries[3]].copy())
            sheet[boundaries[0]: boundaries[2],
                  boundaries[1]: boundaries[3]] = None

    return tables


def find_tables_by_name(tables, name):
    """
    Looks for tables based on a regex expression representing their name.

    :arg np.array: the tables.
    :arg str: the name, in regex.
    :returns List<np.array>: the tables found.
    """
    named_tables = []
    for table in tables:
        if re.match(name, table[0, 0], re.IGNORECASE):
            named_tables.append(table)
    return named_tables


def find_glossary(tables: List[np.array]) -> Dict[str, np.array]:
    """
    Locates the glossarytables, and places them in a dictionary for each type.
    The five tables it looks for are:

        * Type
        * Function
        * Constant
        * Relation
        * Boolean

    :returns Dict[str, np.array]: containing the glossary for type, function,
    constant, relation and boolean.
    """
    glossary = {"Type": None, "Function": None, "Constant": None,
                "Relation": None, "Boolean": None}

    glossary["Type"] = find_glossary_table(tables, "type", critical=True)
    glossary["Function"] = find_glossary_table(tables, "function")
    glossary["Constant"] = find_glossary_table(tables, "constant")
    glossary["Relation"] = find_glossary_table(tables, "relation")
    glossary["Boolean"] = find_glossary_table(tables, "boolean")

    return glossary


def find_glossary_table(tables: List[np.array], name: str,
                        critical: bool = False) -> np.array:
    """
    Looks for a specific glossary table with name "name".

    If the critical boolean is set, an error is returned when no glossary is
    found. For example, there should always be a type glossary.
    Non-critical glossaries only print warnings when none are found.

    :arg tables: the list of arrays, each containing a table.
    :arg name: the name of the table to find.
    :arg critical: True if a table needs to be found. E.g. if the table is not
        found, an error is thrown.
    :returns np.array: the glossary table, if found.
    """
    glossaries = find_tables_by_name(tables, name)
    if len(glossaries) == 0:
        if critical:
            raise ValueError(f"No {name} glossary table was found.")
        else:
            print(f"INFO: No {name} glossary table found.")
        return None

    if len(glossaries) > 1:
        raise ValueError(f"Multiple {name} glossary tables were found.")
    return glossaries[0]


def find_datatables(tables: List[np.array]) -> np.array:
    """
    Locates the Datatable, which contains data that needs to be expressed in
    the structure or our idp file.

    :arg tables: the list of arrays, each containing a table.
    :returns np.array: containing the datatable
    """
    datatables = find_tables_by_name(tables, r"(?i)DataTable|Data Table")
    return datatables


def find_execute_method(tables: List[np.array]) -> Dict[str, object]:
    """
    Locates the Goal table, which contains the inference method.
    There are four possible inference methods:

        * "get x models", where x is the number of required models;
        * "get all models", in order to get all models;
        * "minimize x", where x is a term to minimize;
        * "maximize x", where x is a term to maximize.

    :arg tables: the list of arrays, each containing a table.
    :returns Dict[str, str]: contains the inference method and its applicable
        variables.
    """
    goal_table = find_tables_by_name(tables, 'Goal')
    inf = {"method": "mx",
           "nbmodels": 1,
           "term": ""}
    if len(goal_table) > 1:
        raise ValueError("Only one Goal table allowed")
    # Set the inference method based on the Goal table's content.
    if goal_table:
        cell = str(goal_table[0][1])

        # Check for modelexpansion, minimization and maximization.
        if re.search("(?i) Model*", cell):
            # Extract the amount of models.
            if re.search("all", cell):
                nb = 0
            else:
                nb = re.findall('[0-9]+', cell)[0]
            inf["nbmodels"] = int(nb)
        elif re.search("(?i)Minimize", cell):
            inf["method"] = "min"
            term = re.search(r"(?i)(?<=Minimize) (.*)\'\]$", cell).groups()[0]
            inf["term"] = term
        elif re.search("(?i)Maximize", cell):
            inf["method"] = "max"
            term = re.search(r"(?i)(?<=Maximize) (.*)\'\]$", cell).groups()[0]
            inf["term"] = term
    return inf


def find_tables(tables: List[np.array]) -> List[np.array]:
    """
    Looks for decision tables and constraint tables.

    :arg tables: the list of arrays, each containing a table.
    :returns List[np.array]: a list containing only the decision and constraint
        tables.
    """
    tables = list(filter(lambda table:
                         not re.match('Glossary|DataTable|Comment|Data Table',
                                      table[0, 0],
                                      re.IGNORECASE),
                         tables))
    return tables


def replace_with_error_check(tables: List[np.array], check_error: str,
                             target_error, rule_error=None, deps=[]
                             ) -> List[np.array]:
    """
    Replaces a model by one for error checking.
    There are three types of errors which can be checked for errors:

        * Overlap in the inputs (multiple rules fire for same input).
        * Gaps in the inputs (input for which no rules applies).
        * Shadowed rules (rules which can never fire).

    :arg tables: the list of arrays, each containing a table.
    :arg target_error: the table which we want to find.
    :arg check_error: the type of error check to do.
    :arg rule_error: the rule to check for errors. Only for shadowed.
    :returns List[np.array]: a list containing only the decision and constraint
    """
    relation = True
    nb_rows = None

    overlap = True if check_error == 'overlap' else False
    shadowed = True if check_error == 'shadowed' else False
    gap = True if check_error == 'gap' else False
    new_tables = []
    for i, table in enumerate(tables):
        # Add the Rule relation to an existing relation table.
        if relation and re.search('Relation', table[0][0]):
            relation = False
            # TODO: add Rule row relation
            continue

        # Add a new relation table if one doesn't already exist.
        elif relation and i == len(tables) - 1:
            relation = False
            relation_table = np.empty([3, 1], dtype=object)
            relation_table[0] = ['Relation']
            relation_table[1] = ['Name']
            relation_table[2] = ['Rule row']

            new_tables.append(relation_table)
            continue

        # Add the `row` type.
        elif re.search('Type', table[0][0]):
            # If we already know the number of rows our table has.
            if nb_rows:
                row = [['row', 'int', f'[1..{nb_rows}]']]
            # If we don't, we need to find out by searching for the target
            # table.
            else:
                for table2 in tables:
                    if re.match(target_error, table2[0][0]):
                        row = table2.shape[0] - 2
            try:
                new_tables.append(np.append(table,  row, axis=0))
            except UnboundLocalError:
                raise IOError("Target table not found. Be sure to input the"
                              " name of the last output variable of the table")
            continue

        # Add the `NbRules` constant, which represents number of fired rules.
        elif re.search('Constant', table[0][0]):
            new_tables.append(np.append(table, [['NbRules', 'int']], axis=0))
            continue

        # Change the output of the table to the `Rule` relation.
        # This output is different between U and A/F tables.
        elif re.match(target_error, table[0][0]):
            hit_policy = table[1][0]

            # Reduce multiple output columns to a singular one.
            try:
                while table[0][-2] is None:
                    table = np.delete(table, -2, 1)
            except IndexError:  # When table is too small.
                continue

            # If the hit_policy is Unique, we add a new output for every
            # possible row (so `Rule 0`, `Rule 1`, ...). Each of these outputs
            # is assigned `Yes` whenever their row fires. In other words, we
            # would append the following matrix to a table with three rows:
            # Rule 0 | Rule 1 | Rule 3
            # Yes    | No     | No
            # No     | Yes    | No
            # No     | No     | Yes
            if hit_policy == "U" or shadowed:
                nb_rows = table.shape[0] - 2
                new_dec_table = None

                # Generate new columns and add them to the table
                for i in range(nb_rows):
                    header = f'Rule {i}'
                    new_col = [None, header] + ['No']*nb_rows
                    new_col[i+2] = 'Yes'
                    print(new_col)
                    if new_dec_table is not None:
                        new_dec_table = np.insert(new_dec_table, -1,
                                                  values=new_col, axis=1)
                    else:
                        new_dec_table = np.insert(table, -1,
                                                  values=new_col, axis=1)
                # Delete the actual output column.
                new_dec_table = np.delete(new_dec_table, -1, 1)

            # If the hit_policy is A or F, we need to set a different row value
            # for every different output to ensure that only overlapping rules
            # with conflicting outputs fire.
            elif hit_policy == "F" or hit_policy == "A":
                row_output = {}
                # Find number of unique outputs
                unique_outputs = set(table.T[-1][2:])
                nb_rows = len(unique_outputs)

                # Create a dict mapping every unique output value on a number.
                row_val = 0
                for value in table.T[-1][2:]:
                    if value not in row_output:
                        row_output[value] = row_val
                        row_val += 1
                # Create an output column for every unique output value.
                # This is similar to the way it works for U tables, except that
                # some rows can share the same predicate.
                new_dec_table = None
                for i in range(len(unique_outputs)):
                    header = f'Rule {i}'
                    new_col = [None, header] + ['No'] * len(table.T[-1][2:])
                    for j, val in enumerate(table.T[-1][2:]):
                        if row_output[val] == i:
                            new_col[j+2] = 'Yes'
                    if new_dec_table is not None:
                        new_dec_table = np.insert(new_dec_table, -1,
                                                  values=new_col, axis=1)
                    else:
                        new_dec_table = np.insert(table, -1,
                                                  values=new_col, axis=1)

                # Delete the actual output column.
                new_dec_table = np.delete(new_dec_table, -1, 1)

            new_tables.append(new_dec_table)
            continue

        else:
            # If a table ends up here, it didn't need to be altered.
            # However, if it is not a dependency of the table which we want to
            # check for errors, we do not need to keep it.
            if table[0][0] in deps:
                new_tables.append(table)

    # For every type of error check we need to add specific tables.
    if overlap:
        # Add a C+ table to count number of fired rules.
        count_table = np.array([['Count', 'Count', 'Count', None],
                                ['C+', 'row', 'Rule row', 'NbRules'],
                                ['1', '-', 'Yes', '1']])
        new_tables.append(count_table)

        # Add a constraint that at least two rules have to fire.
        constr_table = np.array([['Fired rules', None],
                                 ['E*', 'NbRules'],
                                 ['1', '> 1']])
        new_tables.append(constr_table)

    elif shadowed:
        # Add a table stating that a specific rule must fire.
        constr_table = np.array([['Fired rule', None],
                                ['E*', f'Rule {rule_error}'],
                                ['1', 'Yes']])
        new_tables.append(constr_table)

    elif gap:
        # Add a table stating that no rule can fire.
        constr_table = np.array([['Fired rule', 'Fired rule', None],
                                ['E*', 'row', 'Rule row'],
                                ['1', '-', 'No']])
        new_tables.append(constr_table)

    return new_tables


def create_json(glossary) -> dict:
    """
    Function to form the json file.

    :arg glossary: the glossary
    :returns None:


    """
    json_dict = {"title": "Title", "timeout": "20"}
    json_dict["symbols"] = glossary.to_json_dicts()
    json_dict["values"] = []
    return json_dict


def create_voc(glossary, target_lang: str = 'idp'):
    """
    Function which creates the vocabulary for the IDP file.

    :arg glossary:
    :arg target_lang: the target output format. Either `idp` or `idpz3`.
    :returns str: the vocabulary for the IDP file.
    """

    lang_dict = {'idp':  {'voc': '\n\nVocabulary V {\n'},
                 'idpz3': {'voc': '\n\nvocabulary V {\n'}}
    voc = lang_dict[target_lang]['voc']
    voc += glossary.to_idp_voc(target_lang)
    voc += "}\n"
    return voc


def create_main(inf: Dict[str, object], parser: Parser,
                target_lang: str = 'idp') -> str:
    """
    Function which creates the main for the IDP file.

    :arg inf: a dictionary containing the inference method and it's applicable
        variables (e.g. what the term looks like, or how many models should be
        generated)
    :arg parser: the parser, needed to parse the optimization term.
    :arg target_lang: the output format to create. Either idp or idpz3.
    :returns str: the main for the IDP file.
    """
    mainstr = "\n\nprocedure main(){\n"
    if inf["method"] == "mx":
        if target_lang == "idp":
            mainstr += "begin = os.clock()\n"
            mainstr += f"stdoptions.nbmodels = {inf['nbmodels']}\n"
            mainstr += "printmodels(modelexpand(T,S))\n"
        elif target_lang == "idpz3":
            mainstr += f"print(model_expand(T,S,{inf['nbmodels']}))\n"
        else:
            raise ValueError(f"Target lang {target_lang} not known")
    elif inf["method"] == "min" or inf["method"] == "max":
        # Create and add termblock
        # Below line generates something like "term = 1", which is then split.
        # We need to pass 1 as argument otherwise the parser wouldn't recognize
        # the term. It's a dirty workaround but works for now. TODO: fix!
        term = parser.parse_val(inf["term"], 1, None).split('=')[0]
        termstr = "term t:V{\n"
        termstr += f"{term}\n"
        termstr += "}\n\n"

        mainstr += "begin = os.clock()\n"
        if inf["method"] == "min":
            mainstr += "begin = os.clock()\n"
            mainstr += "printmodels(minimize(T,S,t))\n"
        else:
            mainstr += "begin = os.clock()\n"
            mainstr += "printmodels(maximize(T,S,t))\n"
        mainstr = termstr + mainstr
    if target_lang == "idp":
        mainstr += "print(\"Elapsed Time:\")"
        mainstr += "print(os.clock() - begin)"
    mainstr += "}\n"
    return mainstr


def create_struct(tables: List[np.array],
                  parser: Parser,
                  glossary: Glossary,
                  target_lang: str = 'idp') -> str:
    """
    The structure consists of data supplied by data tables.

    The data inside data tables is interpreted and set as "struct_args".
    "struct_args" is then used in the to_idp_struct() method to form the
    structure.

    :arg tables: the list of arrays, each containing a table.
    :arg parser: the parser to read the headers of data tables.
    :arg glossary: the glossary
    :arg target_lang: the output format to create. Either idp or idpz3.
    :returns str: the structure for the IDP file.
    """
    lang_dict = {'idp':  {'struct': '\n\nStructure S:V {\n'},
                 'idpz3': {'struct': '\n\nstructure S: V{\n'}}

    # Interpret the Datatables.
    glossary.read_datatables(find_datatables(tables), parser)

    # Form the structure.
    struct = lang_dict[target_lang]['struct']
    for pred in glossary.predicates + glossary.types:
        s = pred.to_idp_struct(target_lang)
        if s is None:
            continue
        struct += s
    struct += "}\n"

    return struct


def create_theory(tables: List[np.array], parser: Parser,
                  aux_var_needed: bool,
                  target_lang: str = 'idp') -> str:
    """
    Function to create the theory in the IDP format.

    :arg tables: the list of arrays, each containing a table.
    :arg parser: the parser.
    :arg aux_var_needed: a booloan signifying whether auxiliary variables are
        needed. This is the case for C# aggregates, when no interface is used.
    :arg target_lang: the output format for the file. Either `idp` or `idpz3`
    :returns str: the theory for the IDP file.
    """
    lang_dict = {'idp': {'theory': 'Theory T:V {\n'},
                 'idpz3': {'theory': 'theory T: V{\n'}}
    theory = lang_dict[target_lang]['theory']

    inner_t = ""
    for dt in find_tables(tables):
        t = Table(dt, parser, aux_var_needed).export(target_lang)
        if t is None:
            continue
        inner_t += t

    theory += merge_definitions(inner_t)
    theory += "}\n"
    return theory


def find_auxiliary_variables(tables: List[np.array],
                             parser: Parser) -> List[str]:
    """
    Due to a bug in the IDP systems, the variables in a C# table each need to
    use an auxiliary variable to function properly.
    This has the same name as the variable, but preceded with an underscore.

    This function gets the auxiliary variables for every table.
    """
    aux_var: List[str] = []
    for dt in find_tables(tables):
        aux = Table(dt, parser).find_auxiliary()
        if aux:
            aux_var += aux
    return list(filter(None, aux_var))


def create_display(goal_var: List[str]) -> str:
    """
    Format the display block for a DMN specification.
    """
    from cdmn.idpname import idp_name
    display = "\ndisplay { \n"
    for var in goal_var:
        display += f"\t goal(`{idp_name(var)}).\n"

    display += "}\n"
    return display


def create_dependency_graph(tables: List[np.array], parser: Parser) -> Dict:
    """
    Function to create dependency graph of variables.
    Basically, we check for every output of a decision table what the inputs
    are. The outputs depend on these inputs.
    By doing this for every table, we can build a graph of what knowledge
    depends on what other knowledge.

    :arg tables: the list of arrays, each containing a table.
    :arg parser: the parser.
    :returns str: the theory for the IDP file.
    """

    dependency_graph: Dict = {}
    for dt in find_tables(tables):
        t = Table(dt, parser, False)

        # Only decision tables count as dependency.
        if t.hit_policy == "E*":
            continue
        for outp in t.outputs:
            if outp not in dependency_graph:
                dependency_graph[outp] = t.inputs
            else:
                for inp in t.inputs:
                    dependency_graph[outp].append(inp)
        if t is None:
            continue

    return dependency_graph


def get_dependencies(variable, dependency_graph, level=0, downstream=False
                     ) -> Dict:
    """
    If downstream is False, then the function returns all the upstream
    dependencies, i.e., the variables that need to have values before the
    variable's value can be set.

    If downstream is set to True, it does the opposite: it returns all the
    upstream dependencies, i.e. the variables that depend on this variable's
    value.
    """
    # print(variable, level)
    dependencies = {}
    if not downstream:
        if variable not in dependency_graph.keys():
            # The variable has no upstream dependencies.
            pass
        else:
            for dep_var in dependency_graph[variable]:
                # If dep_var is not yet in the dependencies list, add it.
                # If it's already in, but at higher level, add the lower level.
                if (dep_var not in dependencies or
                        level < dependencies[dep_var]):
                    dependencies[dep_var] = level

                for dep, nlevel in get_dependencies(dep_var, dependency_graph,
                                                    level=level+1,
                                                    downstream=False).items():
                    # If the dependency is already present, keep the lowest
                    # one.
                    if dep not in dependencies or nlevel < dependencies[dep]:
                        dependencies[dep] = nlevel
    else:
        for var, deps in dependency_graph.items():
            if variable in deps:
                # Don't add variables with higher levels, same as above.
                if (variable not in dependencies or
                        level < dependencies[variable]):
                    dependencies[variable] = level
                dependencies[var] = level
                for dep, nlevel in get_dependencies(var, dependency_graph,
                                                    level+1, True).items():
                    if dep not in dependencies or nlevel < dependencies[dep]:
                        dependencies[dep] = nlevel
        pass
    return dependencies
