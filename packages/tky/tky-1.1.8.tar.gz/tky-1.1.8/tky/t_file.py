# _*_ coding:utf-8 _*_
"""
常用文件类集合
\n 目前支持方法有：
\n Jenkins_t(Job_num/file_list) （Jenkins类处理）
\n File_val(xml_msg/path_val/last_val) （文件类处理，依赖os/xml.dom）
\n Excel(read_data/write_data) （Excel处理，依赖xlrd/openpyxl）
\n @author: 'TangKaiYue'
"""

import os
from .t_jde import null


# Jenkins类集合
class Jenkins_t:
    """
    【Jenkins类集合】\n
    \n 目前已支持方法：
    \n 1、Job_num | 获取最新构建编号文件夹
    \n 2、file_list | 拼接Jenkins里构建日志的文件路径
    """

    # 获取最新的构建编号文件夹
    def Job_num(self: str = 'null'):
        """
        【获取Jenkins最新的构建编号文件夹】\n
        :param  self : str (job文件夹路径,如：C:\jenkins\Test\jobs\API_Master)；
        :return: 最新构建的文件夹编号；
        """
        try:
            if null(self) == 'N':
                return print('入参异常：job文件夹路径 不可为空...')
            else:
                f = open(str(self) + '\\nextBuildNumber', 'r')
                num = int(f.read()) - 1
                return num
        except EnvironmentError as err:
            return '发生异常：%s' % err

    # 拼接Jenkins构建日志文件的路径
    def file_list(self: str = 'null', file: str = 'null'):
        """
        【拼接Jenkins构建日志相关文件路径】\n
        :param  self : str (job文件夹路径,如：C:\jenkins\Test\jobs\API_Master)；
        :param  file : str (文件名称,如：build.xml)；
        :return: 拼接的文件路径；
        """
        try:
            if null(self) == 'N' or null(file) == 'N':
                return print('入参异常：job文件夹路径、文件名称 不可为空...')
            else:
                f = open(str(self) + '\\nextBuildNumber', 'r')
                num = int(f.read()) - 1
                path = str(self) + '\\builds\\' + str(num) + "\\" + str(file)
                return str(path)
        except EnvironmentError as err:
            return '发生异常：%s' % err


# 文件处理集合
class File_val:
    """
    【文件处理集合】\n
    \n 目前已支持方法：
    \n 1、xml_msg | 读取xml文件相应数据(依赖xml.dom)
    \n 2、path_val | 取文件夹内某个文件
    \n 3、last_val | 取某文件某行内容
    """

    # 打开xml文件，并获取对应数据
    def xml_msg(self: str = 'null', path: str = 'null'):
        """
        【打开xml文件，并获取对应数据】\n
        :param  self : str 要获取的数据组，如：['1','2']；
        :param  path : str 要打开的xml文件路径；
        :return 以list输出获取的数据；
        """
        try:
            from xml.dom import minidom  # 导入xml模块
            if null(self) == 'N' or null(path) == 'N':
                return '入参异常：请传入不为空的 self,path 值...'
            else:
                dom = minidom.parse(path)
                xml_list = []
                for e in self:
                    result = dom.getElementsByTagName(str(e))  # 获取操作结果
                    for i in range(len(result)):
                        if result[i].firstChild is None:
                            return "null"
                        else:
                            result1 = result[i].firstChild.data  # 获取节点数据
                            xml_list.append(result1)
                return xml_list
        except EnvironmentError as err:
            return '发生异常：%s' % err

    # 查询文件夹内的某个文件
    def path_val(self: str = 'null', val: int = 'null'):
        """
        【查询文件夹内的某个文件】\n
        :param  self : str 文件夹路径,如：C:\Test；
        :param  val : int  哪个文件,如最后一个：-1；不传查全部；
        :return  文件的名称；
        """
        try:
            if null(self) == 'N':
                return '入参异常：请传入不为空的 文件夹路径 值...'
            else:
                if null(val) == 'N':
                    listdir = os.listdir(self)  # 取文件夹所有
                    return str(listdir)
                else:
                    listdir = os.listdir(self)
                    val_name = listdir[int(val)]  # 取某个文件
                    return str(val_name)
        except EnvironmentError as err:
            return '发生异常：%s' % err

    # 查询某文件的某行内容
    def last_val(self: str = 'null', val: int = 'null'):
        """
        【查询某文件的某行内容】\n
        :param  self : str 文件路径；
        :param  val : int  行数,如最后一行：-1；未传取所有；
        :return  获取的最后一行内容；
        """
        try:
            if null(self) == 'N':
                return '入参异常：请传入不为空的 路径 值...'
            else:
                if null(val) == 'N':
                    with open(self, 'r', encoding='utf-8') as f:  # 打开文件
                        lines = f.readlines()  # 读取所有行
                        return str(lines)
                else:
                    with open(self, 'r', encoding='utf-8') as f:  # 打开文件
                        lines = f.readlines()  # 读取所有行
                        last_line = lines[int(val)]  # 取某一行
                        return str(last_line)
        except EnvironmentError as err:
            return '发生异常：%s' % err


# Excel处理集合
class Excel:
    """
    【Excel处理集合】\n
    \n 目前已支持方法：
    \n 1、read_data | 读取Excel表格内容(依赖xlrd)
    \n 2、write_data | 写入Excel内容(依赖openpyxl)
    """

    # 读取Excel内容
    def read_data(self: str = 'null', stn: str = "Sheet1"):
        """
        【读取Excel表格内容】\n
        :param self: str 文件路径
        :param stn: str 工作表,默认为Sheet1
        :return: 读取结果
        """
        try:
            if null(self) == 'N':
                return print('入参异常：文件路径 不可为空...')
            else:
                import xlrd
                data = xlrd.open_workbook(self)
                table = data.sheet_by_name(stn)
                # 获取总行数、总列数
                nrows = table.nrows
                ncols = table.ncols
                print('%s 工作表数据概况：共计%d行，%d列' % (stn, nrows, ncols))
                if nrows > 0:
                    # 获取第一行的内容，列表格式
                    keys = table.row_values(0)
                    list_data = []
                    # 获取每一行的内容，列表格式
                    for col in range(0, nrows):
                        values = table.row_values(col)
                        # keys，values组合转换为字典
                        val_dict = dict(zip(keys, values))
                        list_data.append(val_dict)
                    # print(list_data)
                    return list_data
                else:
                    print("表格是空数据!")
                    return None
        except EnvironmentError as err:
            return '发生异常：%s' % err

    # 写入Excel内容
    def write_data(self, stn: str = "Sheet1", row_n: int = "null", dow_n: int = "null", value: str = "null"):
        """
        【写入Excel内容】(依赖openpyxl)\n
        :param self: str 写入数据的文件路径
        :param row_n: int 所在行数,单数
        :param dow_n: int 所在列数,单数
        :param value: int 写入内容
        :param stn: str 指定工作表,默认Sheet1
        :return: 无
        """
        try:
            if not os.path.exists(self):
                return print('入参异常：文件路径异常，请检查...')
            else:
                if null(row_n) == 'N' or null(dow_n) == 'N':
                    return print('入参异常：未指定表格的行和列...')
                else:
                    from openpyxl import load_workbook
                    wb = load_workbook(self)  # 获取工作薄-写入的文件
                    ws = wb.get_sheet_by_name(stn)  # 获取工作表1-指定写入的表
                    ws.cell(row_n, dow_n, value)  # 写入内容
                    wb.save(self)
                    return print('%d行 %d列的 %s 内容写入成功！' % (row_n, dow_n, value))
        except EnvironmentError as err:
            return '发生异常：%s' % err
