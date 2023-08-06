import xlrd, xlwt, openpyxl
from xlrd import xldate_as_tuple

from openpyxl import load_workbook
import datetime

'''
xlrd中单元格的数据类型
数字一律按浮点型输出，日期输出成一串小数，布尔型输出0或1，所以我们必须在程序中做判断处理转换
成我们想要的数据类型
0 empty,1 string, 2 number, 3 date, 4 boolean, 5 error
'''


class ExcelData():
    # 初始化方法
    def __init__(self, data_path, sheetname):
        # 定义一个属性接收文件路径
        self.data_path = data_path
        # 定义一个属性接收工作表名称
        self.sheetname = sheetname
        # 使用xlrd模块打开excel表读取数据
        self.data = xlrd.open_workbook(self.data_path)
        # 根据工作表的名称获取工作表中的内容（方式①）
        self.table = self.data.sheet_by_name(self.sheetname)
        # self.table,_ = self.get_sheet()
        # 根据工作表的索引获取工作表的内容（方式②）
        # self.table = self.data.sheet_by_name(0)
        # 获取第一行所有内容,如果括号中1就是第二行，这点跟列表索引类似
        self.keys = self.table.row_values(0) if self.table and self.table.nrows > 0 else []
        # 获取工作表的有效行数
        self.rowNum = self.table.nrows
        # 获取工作表的有效列数
        self.colNum = self.table.ncols

    # 定义一个读取excel表的方法
    def readExcel(self):
        # 定义一个空列表
        datas = []
        for i in range(1, self.rowNum):
            # 定义一个空字典
            sheet_data = {}
            for j in range(self.colNum):
                # 获取单元格数据类型
                c_type = self.table.cell(i, j).ctype
                # 获取单元格数据
                c_cell = self.table.cell_value(i, j)
                if c_type == 2 and c_cell % 1 == 0:  # 如果是整形
                    c_cell = int(c_cell)
                elif c_type == 3:
                    # 转成datetime对象
                    date = datetime.datetime(*xldate_as_tuple(c_cell, 0))
                    c_cell = date.strftime('%Y/%m/%d %H:%M:%S')
                elif c_type == 4:
                    c_cell = True if c_cell == 1 else False
                sheet_data[self.keys[j]] = c_cell
                # 循环每一个有效的单元格，将字段与值对应存储到字典中
                # 字典的key就是excel表中每列第一行的字段
                # sheet_data[self.keys[j]] = self.table.row_values(i)[j]
            # 再将字典追加到列表中
            datas.append(sheet_data)
        # 返回从excel中获取到的数据：以列表存字典的形式返回
        return datas

    def write_cell(self, row_num, col_num, value):
        sheet, wb = self.get_sheet()
        sheet.cell(row_num, col_num, value)  # 行，列，值 这里是从1开始计数的
        wb.save(self.data_path)  # 一定要保存

    def get_sheet(self):
        wb = load_workbook(filename=self.data_path)
        if not wb:
            wb = openpyxl.Workbook()  # 新建工作簿
        if self.sheetname not in wb.sheetnames:
            wb.create_sheet(self.sheetname)  # 添加页
            # sheet = wb.active  # 获得当前活跃的工作页，默认为第一个工作页
        sheet = wb.get_sheet_by_name(self.sheetname)  # 获得指定名称页
        return sheet, wb

    def delete_sheet(self):
        ws, wb = self.get_sheet()
        wb.remove(ws)
        wb.save(self.data_path)

    def delete_row(self, row_num):
        wb = load_workbook(filename=self.data_path)
        sheet = wb.get_sheet_by_name(self.sheetname)  # 获得指定名称页
        for row in range(row_num, sheet.max_row):
            for column in range(sheet.max_column):
                sheet[row][column].value = sheet[row + 1][column].value
        for cell in list(sheet.rows)[sheet.max_row - 1]:
            cell.value = None
        wb.save(self.data_path)

    # col_num是从1开始，不是从0开始
    def delete_column(self, col_num):
        wb = load_workbook(filename=self.data_path)
        sheet = wb.get_sheet_by_name(self.sheetname)  # 获得指定名称页
        sheet.delete_cols(col_num)
        wb.save(self.data_path)

    # 'A1:G37'
    def delete_mult_cell(self, area='A1:A1'):
        ws, wb = self.get_sheet()
        for row in ws[area]:
            for cell in row:
                cell.value = None
        wb.save(self.data_path)

    # 追加一行，values为字符串数组
    def append_row(self, values):
        sheet, wb = self.get_sheet()
        # items = [i for i in values]
        # for row in range(0, len(values)):
        #     sheet.append([row])
        sheet.append(values)
        wb.save(self.data_path)

    # 写入列
    def write_column(self, values=[], col_num=0):
        sheet, wb = self.get_sheet()
        rows = sheet.max_row
        columns = sheet.max_column
        col_num = columns + 1 if col_num <= 0 else col_num
        # column_data = []
        assert len(values) == rows, 'values length: %s must equal with rows length: %s' % (len(values), rows)
        for i in range(1, rows + 1):
            sheet.cell(i, col_num, values[i - 1])  # 行，列，值 这里是从1开始计数的
        wb.save(self.data_path)

    # 获取表格的总行数和总列数
    def get_row_col_num(self):
        sheet, wb = self.get_sheet()
        rows = sheet.max_row
        columns = sheet.max_column
        return rows, columns

    # 获取某个单元格的值
    def get_cell_value(self, row, column):
        sheet, wb = self.get_sheet()
        cell_value = sheet.cell(row=row, column=column).value
        return cell_value

    # 获取某列的所有值
    def get_col_value(self, column):
        sheet, wb = self.get_sheet()
        rows = sheet.max_row
        column_data = []
        for i in range(1, rows + 1):
            cell_value = sheet.cell(row=i, column=column).value
            column_data.append(cell_value)
        return column_data

    # 获取某列的所有值
    def get_col_obj(self, column):
        sheet, wb = self.get_sheet()
        rows = sheet.max_row
        column_data = []
        for i in range(1, rows + 1):
            cell_value = sheet.cell(row=i, column=column)
            column_data.append(cell_value)
        return column_data

    # 获取列超链接
    def get_col_hyberlink(self, column):
        col = excel.get_col_obj(column)
        ret = []
        for c in col:
            link = None
            if c.hyperlink:
                # print(c.hyperlink.display)
                link = c.hyperlink.display
            ret.append(link)
        return ret

    # 获取某行所有值
    def get_row_value(self, row):
        sheet, wb = self.get_sheet()
        columns = sheet.max_column
        row_data = []
        for i in range(1, columns + 1):
            cell_value = sheet.cell(row=row, column=i).value
            row_data.append(cell_value)
        return row_data

    def copy_sheet_from(self, fiel_src, ws_src_name):
        wb_src = load_workbook(filename=fiel_src)
        ws_src = wb_src[ws_src_name]

        sheet, wb = self.get_sheet()

        # 两个for循环遍历整个excel的单元格内容
        for i, row in enumerate(ws_src.iter_rows()):
            for j, cell in enumerate(row):
                sheet.cell(row=i + 1, column=j + 1, value=cell.value)

        wb.save(self.data_path)


def read_to_dict_list(data_path, sheetname='Sheet1'):
    get_data = ExcelData(data_path, sheetname)
    datas = get_data.readExcel()
    return datas


if __name__ == "__main__":
    data_path = "/Users/jinlong.rhj/Book1.xlsx"
    sheetname = "Sheet1"
    excel = ExcelData(data_path, sheetname)
    # datas = get_data.readExcel()
    # print(datas)
    links = excel.get_col_hyberlink(2)
    excel.write_column(links)
