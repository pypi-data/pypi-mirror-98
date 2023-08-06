import numpy as np
from pandas_flavor import register_dataframe_accessor

from openpyxl import Workbook
from openpyxl.utils.dataframe import dataframe_to_rows
from openpyxl.styles import PatternFill, Border, Side, Alignment, Font, NamedStyle
from openpyxl.drawing.image import Image
from openpyxl.utils import get_column_letter

default_col_width = 15
default_font_style = "微软雅黑"

blue = "497FE7"
orange = "F49E5E"
green = "B4EEE3"
yellow = "F5F8AD"


def set_font_format(
    font_size=12,
    bold=False,
    italic=False,
    font_horizontal="center",
    fill_border=True,
    fill_bgcolor=None,
    font_color="black",
    style_name="sty1",
    number_format="General",
):
    """
    set style for font and cell.
    style_name is a required parameter,
    """
    if font_color == "black":
        font_color = "000000"
    elif font_color == "blue":
        font_color = "0000ff"
    elif font_color == "red":
        font_color = "ff4000"
    elif font_color == "green":
        font_color = "33cc33"
    else:
        font_color = "000000"
    fgColor = "C0C0C0" if fill_bgcolor is None else fill_bgcolor
    line = Side(style="thin", color="000000")  # set thin border
    border = Border(top=line, bottom=line, left=line, right=line)
    font = Font(default_font_style, size=font_size, bold=bold, italic=italic, color=font_color)
    fill = PatternFill("solid", fgColor=fgColor)
    alignment = Alignment(
        horizontal=font_horizontal, vertical="center", wrapText=True
    )  # wrapText 自动换行
    if fill_bgcolor is not None:
        style = NamedStyle(
            name=style_name,
            font=font,
            # border=border,
            fill=fill,
            alignment=alignment,
            number_format=number_format,
        )
        if fill_border:
            style.border = border
    else:
        style = NamedStyle(
            name=style_name,
            font=font,
            # border=border,
            alignment=alignment,
            number_format=number_format,
        )
        if fill_border:
            style.border = border
    return style


def merge_series(df, col_name, offset=0, header=True):
    # python list start at 0, so add 1 to match excel
    # add another 1 to account for dataframe header
    default_offset = 1
    if header:
        default_offset += 1
    start_idx = [
        i + default_offset + offset
        for i in df.index[df[col_name].ne(df[col_name].shift())].to_list()
    ]
    max_idx = df.shape[0] + offset + 1
    end_idx = [val - 1 for val in start_idx[1:]]
    if end_idx[-1] < max_idx:
        end_idx.append(max_idx)
    return list(zip(start_idx, end_idx))


@register_dataframe_accessor("excel_format")
class Framer:
    def __init__(self, df, col_offset=0, row_offset=0, workbook=None, worksheet=None):
        self.df = df
        self.workbook = Workbook() if workbook is None else workbook
        self.worksheet = self.workbook.active if worksheet is None else worksheet
        self.default_col_width = 15
        self.row_offset = 0
        self.col_offset = 0
        self.max_row = self.df.shape[0] + 1
        self.max_col = self.df.shape[1]
        self.has_written = False
        self.has_merge_cell = False

    def _get_column_index(self, col_name):
        # fix python index starts at 0 whereas Excel starts at 1
        return self.df.columns.get_loc(col_name) + 1

    def add_header(
        self,
        title_text="",
        cell_height=25,
        font_size=14,
        bold=True,
        fill_bgcolor=None,
        fill_border=False,
    ):
        max_col_letter = get_column_letter(self.max_col)
        self.worksheet.merge_cells(f"A1:{max_col_letter}1")
        header_cell = self.worksheet["A1"]
        header_cell.value = title_text
        fill_bgcolor = "497FE7" if fill_bgcolor is None else fill_bgcolor
        header_cell.style = set_font_format(
            font_size=font_size,
            bold=bold,
            style_name="header",
            fill_border=fill_border,
            fill_bgcolor=fill_bgcolor,
        )
        self.worksheet.row_dimensions[1].height = 25
        self.row_offset += 1
        self.max_row = self.max_row + 1
        return self

    def column_color(self, col_name, fgcolor, skip_row=None):
        min_row = 1 if skip_row is None else skip_row + 1
        col_idx = self._get_column_index(col_name)
        for row in self.worksheet.iter_rows(
            min_row=min_row, max_row=self.max_row, min_col=col_idx, max_col=col_idx
        ):
            for cell in row:
                cell.fill = PatternFill(fgColor=fgcolor, fill_type="solid")
        return self

    def set_columns_color(self, color_map, skip_row=None):
        for col_name, fgcolor in color_map.items():
            self.column_color(col_name, fgcolor, skip_row)
        return self

    def row_color(self, row_num, fgcolor, skip_col=None, max_col=None):
        min_col = 1 if skip_col is None else skip_col
        max_col = self.max_col if max_col is None else max_col
        for row in self.worksheet.iter_rows(
            min_row=row_num, max_row=row_num, min_col=min_col, max_col=max_col
        ):
            for cell in row:
                cell.fill = PatternFill(fgColor=fgcolor, fill_type="solid")
        return self

    def set_rows_color(self, color_map, skip_col=None, max_col=None):
        for row_num, fgcolor in color_map.items():
            self.row_color(row_num, fgcolor, skip_col=skip_col, max_col=max_col)
        return self

    def column_format(self, col_name, num_format):
        format_dict = {"percent": "0.00%", "float": "0.00", "int": "#,##0"}
        col_idx = self._get_column_index(col_name)
        for row in self.worksheet.iter_rows(
            min_row=1, max_row=self.max_row, min_col=col_idx
        ):
            for cell in row:
                cell.number_format = format_dict.get(num_format, num_format)
        return self

    def set_columns_format(self, column_map):
        for col_name, num_format in column_map.items():
            self.column_format(col_name, num_format)
        return self

    def column_width(self, col_name, col_width):
        col_idx = self._get_column_index(col_name)
        col_letter = get_column_letter(col_idx)
        self.worksheet.column_dimensions[col_letter].width = width
        return self

    def set_column_width(self, col_map):
        for col_name, col_width in col_map.items():
            self.column_width(col_name, col_width)
        return self

    def merge_column_cell(self, col_name):
        if not self.has_written:
            self.data_to_cell()
        merge_points = merge_series(self.df, col_name, self.row_offset)
        col_idx = self._get_column_index(col_name)
        col_letter = get_column_letter(col_idx)
        cell_vals = list(self.df[col_name].unique())
        for i, end_point_tuple in enumerate(merge_points):
            start_point, end_point = end_point_tuple
            self.worksheet.merge_cells(
                f"{col_letter}{start_point}:{col_letter}{end_point}"
            )
            _cell = self.worksheet[f"{col_letter}{start_point}"]
            _cell.value = cell_vals[i]
            _cell.alignment = Alignment(horizontal="center", vertical="center")
        self.has_merge_cell = True
        return self

    def set_row_alignment(self, row_num, alignment):
        for row in self.worksheet.iter_rows(min_row=row_num, max_row=row_num):
            for cell in row:
                cell.alignment = Alignment(horizontal=alignment)
        return self

    def data_to_cell(self, set_grid=True):
        header_row_num = self.row_offset + 1
        for row_idx, row in enumerate(
            dataframe_to_rows(self.df, index=False, header=True)
        ):
            for col_idx, value in enumerate(row):
                row_num = row_idx + self.row_offset + 1
                col_num = col_idx + self.col_offset + 1
                self.worksheet.cell(
                    row=row_num,
                    column=col_num,
                ).value = value
                if row_num == header_row_num:
                    self.worksheet.cell(
                        row=row_num, column=col_num
                    ).alignment = Alignment(horizontal="center")
                    self.worksheet.cell(row=row_num, column=col_num).font = Font(
                        bold=True
                    )
                if set_grid:
                    line = Side(style="thin", color="000000")
                    border = Border(top=line, bottom=line, left=line, right=line)
                    self.worksheet.cell(row=row_num, column=col_num).border = border
        self.has_written = True
        return self

    def save(self, path):
        if not self.has_written:
            self.data_to_cell()
        self.workbook.save(path)
        self.workbook.close()